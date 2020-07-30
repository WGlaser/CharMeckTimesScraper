import requests # for making standard html requests
from bs4 import BeautifulSoup # magical tool for parsing html data

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import GREEN

#pull info from Mecklenburg times
page = requests.get("https://publicnotices.mecktimes.com/search/results.aspx?stid=53&PublicationStates=NC")
soup = BeautifulSoup(page.text, 'html.parser')
property_list = soup.find('table').find_all('tr')

book = Workbook()
sheet = book.active

#header row (with extras!!!)
#AdID, Latitude Longitude St. No, Street, City, Zip, County, Auction, Posted, Square Footage, Tax Value, Bedrooms,
#Bathrooms, School Disctrict, Minimum Bid Amount. 16 total headers
headers = property_list[0]
headerTextList = []
header_content = headers.contents
for i in header_content[2:]: #first two items are blank  (checkmark column), start at third item
	if(i.name == 'th'): #div tag, avoids extraneous extra parts of header_content
		print(i.get_text()) #if it's a header (th) get the text
		headerTextList.append(i.get_text())
headerTextList.append("TaxPin") #info we're going to pull from ad (need link to ad, then use BS)


#add header rowers to Excel file
i = 0;

for row in sheet.iter_rows(min_row=1, max_col=len(headerTextList), max_row=1):
	for cell in row:
		cell.value = headerTextList[i];
		cell.fill = PatternFill(fgColor=GREEN, fill_type = "solid")
		i= i+1;



		


#now for a property item
rowCount = 2

justPropList = property_list[1:]
for prop in justPropList:
	property_content = prop.contents
	x = 0
	propertyData=[] #create blank property data list
	for i in property_content[2:]: #first two items are blank, start at third item
		if(i.name == 'td'): #div tag, avoids extranues extra parts of header_content
			#print(i.get_text()) 
			#print("~~~~")
			propertyData.append(i.get_text()) #append to propetty list
	#print(propertyData)
	for row in sheet.iter_rows(min_row=rowCount, max_col=len(propertyData), max_row=rowCount):
		for cell in row:
			print(propertyData[x])
			cell.value = propertyData[x]
			cell.fill = PatternFill(fgColor=GREEN, fill_type = "solid")
			x = x+1
		rowCount = rowCount+1

	
#now let's get ADId, append to "https://publicnotices.mecktimes.com/search/detail.aspx?detail="
#for indidivual property pages, to then grab taxPin
#UPDATE: abandon below. Not every individual property page has tax pin, and they are in different places
#when they do. BOOOO :(

# for row in sheet.iter_cols(min_row=2, min_col=1, max_row=sheet.max_row, max_col=1):
# 	for cell in row:
# 		propURL = "https://publicnotices.mecktimes.com/search/detail.aspx?detail=" + cell.value
# 		propPage = requests.get(propURL)
# 		propSoup = BeautifulSoup(propPage.text, 'html.parser')
# 		mydivs = propSoup.findAll("p", {"class": "columns2"})
# 		print(mydivs)

book.save("RETest.xlsx")
